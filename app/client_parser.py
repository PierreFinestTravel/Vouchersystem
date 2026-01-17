"""Client file parser module.

This module parses client files to extract traveller names:
- SINGLE mode: Extract names from a .docx confirmation file
- GROUP mode: Extract names and room assignments from a .xlsx booking sheet
"""
import logging
import re
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field

from openpyxl import load_workbook
from docx import Document

logger = logging.getLogger(__name__)


@dataclass
class RoomGroup:
    """Represents a room with its occupants."""
    room_number: int
    occupants: List[str] = field(default_factory=list)
    
    def get_names_display(self) -> str:
        """Get formatted names for voucher display."""
        return " & ".join(self.occupants)
    
    def get_filename_safe(self) -> str:
        """Get safe filename from names."""
        names = "_".join(self.occupants)
        # Remove special characters
        safe = "".join(c for c in names if c.isalnum() or c in (' ', '-', '_', '&'))
        return safe.replace(' ', '_').replace('&', '_')[:60]


def extract_trip_id(filename: str) -> Optional[str]:
    """Extract 4-digit Trip ID from filename.
    
    Examples:
        '1008 LFA FRM Frilling SA - Orga.xlsx' -> '1008'
        '_1115 BS Vaughan Hawtrey FAO.xlsx' -> '1115'
        'Bestätigung - Thonhauser GM 22122025.docx' -> '1222'
    """
    # Remove path, get just filename
    filename = filename.replace('\\', '/').split('/')[-1]
    
    # Pattern 1: 4 digits at the start (possibly after underscore)
    match = re.match(r'^_?(\d{4})\s', filename)
    if match:
        return match.group(1)
    
    # Pattern 2: 8-digit date at end (DDMMYYYY) - extract MMDD as Trip ID
    match = re.search(r'(\d{2})(\d{2})(\d{4})\.docx$', filename, re.IGNORECASE)
    if match:
        day, month, year = match.groups()
        return f"{month}{day}"  # e.g., 22122025 -> 1222
    
    # Pattern 3: Any 4-digit sequence at the start
    match = re.search(r'(\d{4})', filename)
    if match:
        return match.group(1)
    
    return None


def validate_trip_ids(orga_filename: str, client_filename: str) -> Tuple[bool, str, str]:
    """Validate that ORGA and client file have matching Trip IDs.
    
    Returns: (is_valid, orga_trip_id, client_trip_id)
    """
    orga_id = extract_trip_id(orga_filename)
    client_id = extract_trip_id(client_filename)
    
    if not orga_id:
        logger.warning(f"Could not extract Trip ID from ORGA: {orga_filename}")
    if not client_id:
        logger.warning(f"Could not extract Trip ID from client file: {client_filename}")
    
    is_valid = orga_id == client_id if (orga_id and client_id) else False
    return is_valid, orga_id or "?", client_id or "?"


def parse_single_client_file(file_path: str) -> List[str]:
    """Parse a SINGLE client .docx file to extract traveller names.
    
    Names are ONLY extracted from the client file, never guessed.
    
    Looks for patterns like:
    - "Kundennamen: Thomas & Petra Thonhauser"  (names on same line)
    - "Kundennamen:" followed by names on next lines
    - "Traveller names: Mr John Smith"
    
    Returns:
        List of traveller names extracted from the file.
        Returns empty list if no names found (caller should handle as error).
    """
    doc = Document(file_path)
    names = []
    paragraphs = [p.text.strip() for p in doc.paragraphs]
    
    # Patterns to look for - MUST match one of these
    # Order matters: most specific patterns first
    name_patterns = [
        r'Kundennamen?:\s*(.+)',          # German: Customer name(s)
        r'Traveller\s*names?:\s*(.+)',    # English: Traveller name(s)
        r'Client\s*names?:\s*(.+)',       # English: Client name(s)
        r'Guest\s*names?:\s*(.+)',        # English: Guest name(s)
        r'Reisende[nr]?:\s*(.+)',         # German: Traveller(s)
        r'Gast(?:name)?:\s*(.+)',         # German: Guest/Guest name
        r'Teilnehmer:\s*(.+)',            # German: Participant
    ]
    
    # Header patterns (label on its own line, names on following lines)
    header_patterns = [
        r'^Kundennamen?:?\s*$',            # German: Customer name(s) - alone on line
        r'^Traveller\s*names?:?\s*$',      # English: Traveller name(s)
        r'^Client\s*names?:?\s*$',         # English: Client name(s)
        r'^Guest\s*names?:?\s*$',          # English: Guest name(s)
        r'^Reisende[nr]?:?\s*$',           # German: Traveller(s)
        r'^Teilnehmer:?\s*$',              # German: Participant
    ]
    
    # NOTE: Do NOT use generic "Name:" pattern - it matches company names like "Firmen Name:"
    
    for i, text in enumerate(paragraphs):
        if not text:
            continue
        
        # First, check if names are on the SAME line as the label
        for pattern in name_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                raw_names = match.group(1).strip()
                # Skip if the match is just numbers or too short
                if len(raw_names) < 3 or raw_names.isdigit():
                    continue
                # Clean up and split names
                names_found = parse_name_string(raw_names)
                if names_found:
                    # Validate names are not empty strings
                    valid_names = [n for n in names_found if n and len(n.strip()) >= 2]
                    if valid_names:
                        names.extend(valid_names)
                        logger.info(f"Found names in SINGLE file (same line): {valid_names}")
                        return names  # Return first valid match
        
        # Second, check if this is a header line with names on FOLLOWING lines
        for pattern in header_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                # Found header, look for names in following paragraphs
                logger.info(f"Found name header at line {i}: '{text}'")
                found_names = []
                
                # Read following lines until we hit an empty line or non-name content
                for j in range(i + 1, min(i + 10, len(paragraphs))):
                    next_text = paragraphs[j].strip()
                    
                    if not next_text:
                        # Empty line - stop if we already have names
                        if found_names:
                            break
                        continue
                    
                    # Stop if we hit a new section/label
                    if ':' in next_text and any(kw in next_text.lower() for kw in 
                        ['firmen', 'typ', 'datum', 'link', 'b&b', 'übernachtung', 
                         'geschäftsbedingungen', 'storno', 'einreise', 'impf']):
                        break
                    
                    # Check if this looks like a name (starts with Herr/Frau or has capital letter)
                    if (next_text.startswith(('Herr ', 'Frau ', 'Mr ', 'Mrs ', 'Ms ', 'Dr ')) or
                        (next_text[0].isupper() and len(next_text.split()) >= 2)):
                        # Clean the name - remove annotations like (EZ), (DZ)
                        clean_name = re.sub(r'\s*\([^)]*\)\s*$', '', next_text).strip()
                        # Remove Herr/Frau prefix for cleaner output
                        clean_name = re.sub(r'^(Herr|Frau|Mr\.?|Mrs\.?|Ms\.?|Dr\.?)\s+', '', clean_name).strip()
                        if clean_name and len(clean_name) >= 2:
                            found_names.append(clean_name)
                            logger.info(f"  Found name: '{clean_name}'")
                
                if found_names:
                    logger.info(f"Found names in SINGLE file (multi-line): {found_names}")
                    return found_names
    
    # NO GUESSING - if no pattern matched, return empty list
    # The caller MUST handle this as an error
    logger.error("SINGLE client file parsing FAILED: No name pattern found. "
                 "File must contain 'Kundennamen:', 'Traveller names:', or similar pattern.")
    return []  # Empty = parsing failed


def parse_name_string(raw_names: str) -> List[str]:
    """Parse a name string into individual names.
    
    Handles formats like:
    - "Thomas & Petra Thonhauser"
    - "Mr John Smith & Mrs Jane Smith"
    - "John Smith, Jane Smith"
    """
    names = []
    
    # Check if it's "FirstName & FirstName LastName" format
    if " & " in raw_names and raw_names.count(" ") <= 4:
        parts = raw_names.split(" & ")
        if len(parts) == 2:
            # Try to extract shared last name
            last_part = parts[1].strip().split()
            if len(last_part) >= 2:
                last_name = last_part[-1]
                first_part = parts[0].strip()
                # Check if first part doesn't have the last name
                if last_name not in first_part:
                    names.append(f"{first_part} {last_name}")
                    names.append(parts[1].strip())
                    return names
    
    # Split by common delimiters
    if ", " in raw_names:
        names = [n.strip() for n in raw_names.split(", ")]
    elif " & " in raw_names:
        names = [n.strip() for n in raw_names.split(" & ")]
    elif " and " in raw_names.lower():
        names = [n.strip() for n in re.split(r'\s+and\s+', raw_names, flags=re.IGNORECASE)]
    else:
        names = [raw_names.strip()]
    
    # Add title prefixes if missing
    formatted_names = []
    for name in names:
        if name and not name.lower().startswith(('mr', 'mrs', 'ms', 'dr', 'prof')):
            # Could add Mr/Mrs but keeping as-is for flexibility
            formatted_names.append(name)
        else:
            formatted_names.append(name)
    
    return formatted_names


def parse_group_client_file(file_path: str) -> List[RoomGroup]:
    """Parse a GROUP client Excel file to extract room assignments.
    
    Names are ONLY extracted from the client file, never guessed.
    
    Expected structure:
    - Column with "Room" header: Room number
    - Column with "Last Name" header: Last Name
    - Column with "First Name" header: First Name
    
    Rows with Room number start a new room.
    Rows without Room number but with names are sharing with previous room.
    
    Returns:
        List of RoomGroup objects with occupant names.
        Returns empty list if parsing fails (caller should handle as error).
    """
    wb = load_workbook(file_path, data_only=True)
    
    # Try to find the right sheet
    ws = None
    for sheet_name in ['BookingSheet', 'Booking Sheet', 'Clients', 'Teilnehmer']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    
    if ws is None:
        ws = wb.active
    
    logger.info(f"GROUP file parsing - Using sheet: {ws.title}")
    
    rooms: List[RoomGroup] = []
    current_room: Optional[RoomGroup] = None
    
    # Find the header row with "Room", "Last Name", "First Name"
    header_row = None
    room_col = 1
    last_name_col = 5
    first_name_col = 6
    
    for row in range(1, min(30, ws.max_row + 1)):
        cell_a = str(ws.cell(row, 1).value or "").lower()
        if cell_a == "room":
            header_row = row
            # Check for actual column positions
            for col in range(1, 15):
                val = str(ws.cell(row, col).value or "").lower()
                if val == "room":
                    room_col = col
                elif "last" in val and "name" in val:
                    last_name_col = col
                elif "first" in val and "name" in val:
                    first_name_col = col
            break
    
    if header_row is None:
        logger.warning("Could not find header row in GROUP file")
        header_row = 1
    
    logger.info(f"Header row: {header_row}, Room col: {room_col}, Last Name col: {last_name_col}, First Name col: {first_name_col}")
    
    # Parse data rows
    last_room_row = header_row  # Track when we last saw a room number
    
    for row in range(header_row + 1, ws.max_row + 1):
        room_val = ws.cell(row, room_col).value
        last_name = ws.cell(row, last_name_col).value
        first_name = ws.cell(row, first_name_col).value
        
        # Skip empty rows
        if not last_name and not first_name:
            continue
        
        # Skip header-like rows
        if str(last_name).lower() in ['last name', 'nachname', 'arr./dep.']:
            continue
        
        # Skip metadata rows
        if last_name and any(kw in str(last_name).lower() for kw in ['bitte', 'nächte', 'ez', 'dz']):
            continue
        
        # Build full name
        full_name = ""
        if first_name and last_name:
            full_name = f"{first_name} {last_name}".strip()
        elif last_name:
            full_name = str(last_name).strip()
        elif first_name:
            full_name = str(first_name).strip()
        
        if not full_name:
            continue
        
        # Check if this is a new room
        if room_val and str(room_val).strip():
            try:
                room_num = int(room_val)
                # New room
                current_room = RoomGroup(room_number=room_num, occupants=[full_name])
                rooms.append(current_room)
                last_room_row = row
                logger.info(f"New room {room_num}: {full_name}")
            except ValueError:
                # Not a number, might be PRO or other code - skip
                if str(room_val).upper() not in ['PRO', 'ROOM']:
                    logger.debug(f"Skipping non-numeric room: {room_val}")
        else:
            # No room number - only add if this row is close to the last room
            # (sharing the room with previous occupant)
            if current_room is not None and (row - last_room_row) <= 2:
                current_room.occupants.append(full_name)
                logger.info(f"  Adding to room {current_room.room_number}: {full_name}")
            else:
                logger.debug(f"Skipping orphan name (row {row}, last room row {last_room_row}): {full_name}")
    
    # Validate results - NO GUESSING
    if not rooms:
        logger.error("GROUP client file parsing FAILED: No rooms with valid names found. "
                     "File must have columns 'Room', 'Last Name', 'First Name' with data.")
        return []  # Empty = parsing failed
    
    # Remove rooms with empty occupant names
    valid_rooms = []
    for room in rooms:
        # Filter out any empty names
        valid_occupants = [name for name in room.occupants if name and len(name.strip()) >= 2]
        if valid_occupants:
            room.occupants = valid_occupants
            valid_rooms.append(room)
        else:
            logger.warning(f"Room {room.room_number} has no valid occupant names - skipping")
    
    if not valid_rooms:
        logger.error("GROUP client file parsing FAILED: All rooms have empty or invalid names.")
        return []  # Empty = parsing failed
    
    logger.info(f"Successfully parsed {len(valid_rooms)} rooms from GROUP file:")
    for room in valid_rooms:
        logger.info(f"  Room {room.room_number}: {room.occupants}")
    
    return valid_rooms


def get_all_names_from_rooms(rooms: List[RoomGroup]) -> str:
    """Get all unique names from room groups for display."""
    all_names = []
    for room in rooms:
        all_names.extend(room.occupants)
    return ", ".join(all_names)

