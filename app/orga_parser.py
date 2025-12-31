"""ORGA Excel parser module.

This module parses the ORGA Excel file and extracts all service information
(hotels, transfers, activities, restaurants, car rentals, golf) into structured data.
"""
import logging
from datetime import datetime, date
from typing import Optional, List, Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    ParsedORGA, HotelStay, TransferVoucher, TransferLeg,
    ActivityVoucher, ActivityEntry, RestaurantVoucher,
    CarRentalVoucher, GolfVoucher
)
from .supplier_info import get_supplier_info

logger = logging.getLogger(__name__)


class ColumnMapping:
    """Dynamic column mapping detected from ORGA header row."""
    
    def __init__(self):
        # Hotel columns (defaults)
        self.days = 1
        self.day = 2
        self.date = 3
        self.region_city = 4
        self.hotel_supplier = 5
        self.room = 6
        self.board = 7
        self.hotel_notes = 8
        self.hotel_status = 9
        self.hotel_invoice = 10
        
        # Golf columns (defaults for compact format)
        self.golf_supplier = 11
        self.golf_course = 12
        self.tee_time = 13
        self.driving_range = None  # May not exist
        self.golf_cart = 14
        self.rental_set = 15
        self.golf_notes = 16
        self.golf_status = 17
        self.golf_invoice = 18
        
        # Activity columns (defaults for compact format)
        self.activity_supplier = 20
        self.activity_name = 21
        self.activity_time = 22
        self.activity_notes = 23
        self.activity_status = 24
        self.activity_invoice = 25
        
        # Transfer columns (defaults for compact format)
        self.transfer_supplier = 26
        self.transfer_route = 27
        self.service_type = 28
        self.pickup_time = 29
        self.dropoff_time = 30
        self.flight_num = 31
        self.flight_time = 32
        self.travel_time = 33
        self.transfer_notes = 34
        self.transfer_status = 35
        self.transfer_invoice = 36


def detect_columns(ws: Worksheet, header_row: int) -> ColumnMapping:
    """Auto-detect column positions from header row."""
    mapping = ColumnMapping()
    
    # Read all headers
    headers = {}
    for col in range(1, min(60, ws.max_column + 1)):
        val = ws.cell(header_row, col).value
        if val:
            headers[col] = str(val).lower().strip()
    
    logger.info(f"Detecting columns from header row {header_row}, found {len(headers)} headers")
    
    # Track which section we're in based on column order
    # ORGA format: Hotel -> Golf -> Activity -> Transfer
    golf_start = None
    activity_start = None
    transfer_start = None
    
    for col, header in sorted(headers.items()):
        header_lower = header.lower()
        
        # Hotel section (columns 1-10 typically)
        if header_lower == "days":
            mapping.days = col
        elif header_lower == "day":
            mapping.day = col
        elif header_lower == "date":
            mapping.date = col
        elif header_lower in ["region/city", "region", "city"]:
            mapping.region_city = col
        elif "hotel" in header_lower and "supplier" in header_lower:
            mapping.hotel_supplier = col
        elif header_lower == "room":
            mapping.room = col
        elif header_lower == "board":
            mapping.board = col
        
        # Golf section - detect start by "golf supplier"
        elif "golf" in header_lower and "supplier" in header_lower:
            mapping.golf_supplier = col
            golf_start = col
        elif "golf" in header_lower and "course" in header_lower:
            mapping.golf_course = col
        elif "tee" in header_lower and "time" in header_lower:
            mapping.tee_time = col
        elif "driving" in header_lower and "range" in header_lower:
            mapping.driving_range = col
        elif "golf" in header_lower and "cart" in header_lower:
            mapping.golf_cart = col
        elif "rental" in header_lower and "set" in header_lower:
            mapping.rental_set = col
        
        # Activity section - detect by "supplier" after golf section, or "activity"
        elif header_lower == "supplier" and golf_start and not activity_start:
            # First "Supplier" after golf section = Activity Supplier
            mapping.activity_supplier = col
            activity_start = col
        elif header_lower == "activity":
            mapping.activity_name = col
            if not activity_start:
                # Look back for supplier
                for c in range(col - 1, max(0, col - 5), -1):
                    if headers.get(c, "").lower() == "supplier":
                        mapping.activity_supplier = c
                        activity_start = c
                        break
        
        # Transfer section - detect by "supplier" after activity, or "transport/transfers route"
        elif header_lower == "supplier" and activity_start and not transfer_start:
            mapping.transfer_supplier = col
            transfer_start = col
        elif "transport" in header_lower or "transfer" in header_lower and "route" in header_lower:
            mapping.transfer_route = col
            if not transfer_start:
                # Look back for supplier
                for c in range(col - 1, max(0, col - 3), -1):
                    if headers.get(c, "").lower() == "supplier":
                        mapping.transfer_supplier = c
                        transfer_start = c
                        break
        elif header_lower == "service type":
            mapping.service_type = col
        elif "p/up" in header_lower or "pickup" in header_lower:
            mapping.pickup_time = col
        elif "d/off" in header_lower or "dropoff" in header_lower:
            mapping.dropoff_time = col
        elif "flight" in header_lower and "#" in header_lower:
            mapping.flight_num = col
        elif "flight" in header_lower and "time" in header_lower:
            mapping.flight_time = col
        elif "travel" in header_lower and "time" in header_lower:
            mapping.travel_time = col
    
    # Handle "Notes" and "Status" columns - they appear in each section
    # Find them relative to the section starts
    for col, header in sorted(headers.items()):
        header_lower = header.lower()
        
        if header_lower == "notes" or header_lower == "notes ":
            # Determine which section this Notes belongs to
            if transfer_start and col > transfer_start:
                mapping.transfer_notes = col
            elif activity_start and col > activity_start:
                mapping.activity_notes = col
            elif golf_start and col > golf_start:
                mapping.golf_notes = col
            elif col <= 10:  # Hotel section
                mapping.hotel_notes = col
        
        elif header_lower == "status":
            if transfer_start and col > transfer_start:
                mapping.transfer_status = col
            elif activity_start and col > activity_start:
                mapping.activity_status = col
            elif golf_start and col > golf_start:
                mapping.golf_status = col
            elif col <= 10:
                mapping.hotel_status = col
        
        elif "invoice" in header_lower:
            if transfer_start and col > transfer_start:
                mapping.transfer_invoice = col
            elif activity_start and col > activity_start:
                mapping.activity_invoice = col
            elif golf_start and col > golf_start:
                mapping.golf_invoice = col
            elif col <= 10:
                mapping.hotel_invoice = col
    
    # Log detected columns for debugging
    logger.info(f"Column mapping detected:")
    logger.info(f"  Hotel: supplier={mapping.hotel_supplier}, room={mapping.room}, board={mapping.board}, notes={mapping.hotel_notes}")
    logger.info(f"  Golf: supplier={mapping.golf_supplier}, course={mapping.golf_course}, tee_time={mapping.tee_time}")
    logger.info(f"  Activity: supplier={mapping.activity_supplier}, name={mapping.activity_name}, time={mapping.activity_time}")
    logger.info(f"  Transfer: supplier={mapping.transfer_supplier}, route={mapping.transfer_route}, pickup={mapping.pickup_time}")
    
    return mapping


# Legacy column indices for backwards compatibility (will be overridden by detect_columns)
COL_DAYS = 1
COL_DAY = 2
COL_DATE = 3
COL_REGION_CITY = 4
COL_HOTEL_SUPPLIER = 5
COL_ROOM = 6
COL_BOARD = 7
COL_HOTEL_STATUS = 9
COL_HOTEL_NOTES = 8
COL_HOTEL_STATUS2 = 9
COL_HOTEL_INVOICE = 10

# Golf columns (compact format)
COL_GOLF_SUPPLIER = 11
COL_GOLF_COURSE = 12
COL_TEE_TIME = 13
COL_DRIVING_RANGE = 14
COL_GOLF_CART = 15
COL_RENTAL_SET = 16
COL_GOLF_NOTES = 17
COL_GOLF_STATUS = 18
COL_GOLF_INVOICE = 19

# Activity columns (compact format)
COL_ACTIVITY_SUPPLIER = 20
COL_ACTIVITY_NAME = 21
COL_ACTIVITY_TIME = 22
COL_ACTIVITY_NOTES = 23
COL_ACTIVITY_STATUS = 24
COL_ACTIVITY_INVOICE = 25

# Transfer columns (compact format)
COL_TRANSFER_SUPPLIER = 26
COL_TRANSFER_ROUTE = 27
COL_SERVICE_TYPE = 28
COL_PICKUP_TIME = 29
COL_DROPOFF_TIME = 30
COL_FLIGHT_NUM = 31
COL_FLIGHT_TIME = 32
COL_TRAVEL_TIME = 33
COL_TRANSFER_NOTES = 34
COL_TRANSFER_STATUS = 35
COL_TRANSFER_INVOICE = 36


def get_cell_value(ws: Worksheet, row: int, col: int) -> Optional[str]:
    """Get cell value as string, handling None and whitespace."""
    val = ws.cell(row, col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    return str(val).strip() if str(val).strip() else None


def parse_date(val: Any) -> Optional[date]:
    """Parse a date value from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try parsing string formats
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(str(val), "%d.%m.%Y").date()
    except ValueError:
        pass
    return None


def find_header_row(ws: Worksheet) -> int:
    """Find the row containing column headers."""
    for row in range(1, 30):  # Extended range to find headers in different formats
        # Look for the "Days" header in column 1
        val = get_cell_value(ws, row, COL_DAYS)
        if val and str(val).lower() == "days":
            return row
    return 10  # Default based on analysis


def find_data_start_row(ws: Worksheet, header_row: int) -> int:
    """Find the first row with actual data after headers."""
    for row in range(header_row + 1, header_row + 10):
        date_val = get_cell_value(ws, row, COL_DATE)
        if date_val and parse_date(date_val):
            # Skip example rows
            days_val = get_cell_value(ws, row, COL_DAYS)
            if days_val and str(days_val).lower() == "e.g":
                continue
            return row
    return header_row + 2


def is_car_rental_row(route_val: str) -> bool:
    """Check if a transfer row is actually a car rental."""
    if not route_val:
        return False
    route_lower = route_val.lower()
    return "rental car" in route_lower or "group o" in route_lower or "group " in route_lower


def parse_orga(file_path: str) -> ParsedORGA:
    """Parse an ORGA Excel file and extract all service data."""
    logger.info(f"Parsing ORGA file: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    
    # Find the correct sheet - prefer sheets marked as "correct" or with actual data
    ws = None
    orga_sheets = []
    
    for sheet_name in wb.sheetnames:
        if "orga" in sheet_name.lower():
            orga_sheets.append(sheet_name)
            # Prefer sheets with "correct" in the name
            if "correct" in sheet_name.lower():
                ws = wb[sheet_name]
                logger.info(f"Using sheet (marked as correct): {sheet_name}")
                break
    
    # If no "correct" sheet, find one with actual data
    if ws is None and orga_sheets:
        for sheet_name in orga_sheets:
            test_ws = wb[sheet_name]
            # Check if this sheet has data in the supplier columns
            # Try different header row positions (10 or 19)
            for header_row in [10, 19]:
                for data_row in [header_row + 2, header_row + 1]:
                    if data_row <= test_ws.max_row:
                        # Check hotel supplier column (5)
                        hotel = test_ws.cell(data_row, 5).value
                        if hotel and str(hotel).lower() not in ['hotel supplier', 'e.g', 'example']:
                            ws = test_ws
                            logger.info(f"Using sheet (has data): {sheet_name}")
                            break
                if ws:
                    break
            if ws:
                break
        
        # Fall back to first Orga sheet if none had data
        if ws is None:
            ws = wb[orga_sheets[0]]
            logger.info(f"Using sheet (first Orga): {orga_sheets[0]}")
    
    if ws is None:
        ws = wb.active
        logger.info(f"Using active sheet: {ws.title}")
    
    result = ParsedORGA()
    
    # Extract metadata from header rows
    for row in range(1, 10):
        label = get_cell_value(ws, row, 1)
        value = get_cell_value(ws, row, 4)
        if label and value:
            label_lower = label.lower()
            if "lead name" in label_lower:
                result.client_name = str(value)
            elif "pax" in label_lower:
                try:
                    result.pax = int(value)
                except ValueError:
                    pass
            elif "dates" in label_lower:
                result.dates = str(value)
            elif "trip number" in label_lower:
                result.trip_number = str(value)
    
    # Find header and data rows
    header_row = find_header_row(ws)
    data_start = find_data_start_row(ws, header_row)
    logger.info(f"Header row: {header_row}, Data starts: {data_start}")
    
    # Auto-detect column positions from header row
    cols = detect_columns(ws, header_row)
    
    # Collect all data rows using detected column positions
    data_rows = []
    for row in range(data_start, ws.max_row + 1):
        date_val = get_cell_value(ws, row, cols.date)
        current_date = parse_date(date_val)
        
        if current_date is None:
            # Check if this is an "action" or notes row
            col1_val = get_cell_value(ws, row, 1)
            if col1_val and ("action" in str(col1_val).lower() or "book" in str(col1_val).lower()):
                break  # End of data rows
            continue
        
        data_rows.append({
            "row": row,
            "date": current_date,
            "days": get_cell_value(ws, row, cols.days),
            # Hotel
            "region_city": get_cell_value(ws, row, cols.region_city),
            "hotel_supplier": get_cell_value(ws, row, cols.hotel_supplier),
            "room": get_cell_value(ws, row, cols.room),
            "board": get_cell_value(ws, row, cols.board),
            "hotel_status": get_cell_value(ws, row, cols.hotel_status),
            "hotel_notes": get_cell_value(ws, row, cols.hotel_notes),
            # Golf
            "golf_supplier": get_cell_value(ws, row, cols.golf_supplier),
            "golf_course": get_cell_value(ws, row, cols.golf_course),
            "tee_time": get_cell_value(ws, row, cols.tee_time),
            "golf_cart": get_cell_value(ws, row, cols.golf_cart) if cols.golf_cart else None,
            "rental_set": get_cell_value(ws, row, cols.rental_set) if cols.rental_set else None,
            "golf_notes": get_cell_value(ws, row, cols.golf_notes) if cols.golf_notes else None,
            # Activity
            "activity_supplier": get_cell_value(ws, row, cols.activity_supplier),
            "activity_name": get_cell_value(ws, row, cols.activity_name) if cols.activity_name else None,
            "activity_time": get_cell_value(ws, row, cols.activity_time) if cols.activity_time else None,
            "activity_notes": get_cell_value(ws, row, cols.activity_notes) if cols.activity_notes else None,
            # Transfer
            "transfer_supplier": get_cell_value(ws, row, cols.transfer_supplier),
            "transfer_route": get_cell_value(ws, row, cols.transfer_route) if cols.transfer_route else None,
            "service_type": get_cell_value(ws, row, cols.service_type) if cols.service_type else None,
            "pickup_time": get_cell_value(ws, row, cols.pickup_time) if cols.pickup_time else None,
            "dropoff_time": get_cell_value(ws, row, cols.dropoff_time) if cols.dropoff_time else None,
            "flight_num": get_cell_value(ws, row, cols.flight_num) if cols.flight_num else None,
            "flight_time": get_cell_value(ws, row, cols.flight_time) if cols.flight_time else None,
            "transfer_notes": get_cell_value(ws, row, cols.transfer_notes) if cols.transfer_notes else None,
            "transfer_status": get_cell_value(ws, row, cols.transfer_status) if cols.transfer_status else None,
        })
    
    logger.info(f"Found {len(data_rows)} data rows")
    
    # Log first data row for debugging
    if data_rows:
        first = data_rows[0]
        logger.info(f"First data row sample - Hotel: {first.get('hotel_supplier')}, Activity: {first.get('activity_supplier')}, Transfer: {first.get('transfer_supplier')}")
    
    # Parse hotels - group consecutive stays at the same hotel
    result.hotels = parse_hotels(data_rows)
    logger.info(f"Parsed {len(result.hotels)} hotel stays")
    
    # Parse transfers - group by supplier
    result.transfers = parse_transfers(data_rows)
    logger.info(f"Parsed {len(result.transfers)} transfer vouchers")
    
    # Parse car rentals
    result.car_rentals = parse_car_rentals(data_rows)
    logger.info(f"Parsed {len(result.car_rentals)} car rental vouchers")
    
    # Parse activities - group by supplier
    result.activities = parse_activities(data_rows)
    logger.info(f"Parsed {len(result.activities)} activity vouchers")
    
    # Parse restaurants
    result.restaurants = parse_restaurants(data_rows)
    logger.info(f"Parsed {len(result.restaurants)} restaurant vouchers")
    
    # Parse golf
    result.golf = parse_golf(data_rows)
    logger.info(f"Parsed {len(result.golf)} golf vouchers")
    
    return result


def parse_hotels(data_rows: List[Dict]) -> List[HotelStay]:
    """Parse hotel stays from data rows, grouping consecutive nights."""
    hotels = []
    current_hotel = None
    current_start = None
    
    for i, row in enumerate(data_rows):
        hotel_supplier = row.get("hotel_supplier")
        
        if hotel_supplier:
            # Clean up supplier name (remove trailing whitespace/newlines)
            hotel_supplier = hotel_supplier.strip().split('\n')[0].strip()
            
            if current_hotel is None or hotel_supplier.lower() != current_hotel.lower():
                # New hotel stay - save previous if exists
                if current_hotel is not None:
                    # Find the checkout date (current row's date)
                    checkout = row["date"]
                    nights = (checkout - current_start).days
                    
                    hotels.append(HotelStay(
                        supplier=current_hotel,
                        region_city=current_region,
                        room_type=current_room or "",
                        board=current_board or "",
                        check_in=current_start,
                        check_out=checkout,
                        nights=nights,
                        notes=current_notes or "",
                        status=current_status or ""
                    ))
                
                # Start new hotel
                current_hotel = hotel_supplier
                current_start = row["date"]
                current_region = row.get("region_city", "")
                current_room = row.get("room", "")
                current_board = row.get("board", "")
                current_notes = row.get("hotel_notes", "")
                current_status = row.get("hotel_status", "")
            else:
                # Same hotel - update room/notes if provided
                if row.get("room"):
                    current_room = row.get("room")
                if row.get("hotel_notes"):
                    if current_notes:
                        current_notes += "\n" + row.get("hotel_notes")
                    else:
                        current_notes = row.get("hotel_notes")
    
    # Don't forget the last hotel
    if current_hotel is not None:
        # For the last hotel, checkout is day after the last row
        last_date = data_rows[-1]["date"] if data_rows else current_start
        # Find next day for checkout
        from datetime import timedelta
        checkout = last_date + timedelta(days=1)
        nights = (checkout - current_start).days
        
        hotels.append(HotelStay(
            supplier=current_hotel,
            region_city=current_region,
            room_type=current_room or "",
            board=current_board or "",
            check_in=current_start,
            check_out=checkout,
            nights=nights,
            notes=current_notes or "",
            status=current_status or ""
        ))
    
    # Add supplier info
    for hotel in hotels:
        info = get_supplier_info(hotel.supplier)
        hotel.address = info.get("address", "")
        hotel.phone = info.get("phone", "")
        hotel.gps = info.get("gps", "")
    
    return hotels


def parse_transfers(data_rows: List[Dict]) -> List[TransferVoucher]:
    """Parse transfers from data rows, grouping by supplier."""
    transfers_by_supplier: Dict[str, TransferVoucher] = {}
    
    for row in data_rows:
        supplier = row.get("transfer_supplier")
        route = row.get("transfer_route")
        
        if not supplier:
            continue
        
        # Handle multi-line suppliers/routes
        suppliers = [s.strip() for s in supplier.split('\n') if s.strip()]
        routes = [r.strip() for r in (route or "").split('\n') if r.strip()]
        pickup_times = (row.get("pickup_time") or "").split('\n')
        dropoff_times = (row.get("dropoff_time") or "").split('\n')
        flight_nums = (row.get("flight_num") or "").split('\n')
        
        for idx, sup in enumerate(suppliers):
            rt = routes[idx] if idx < len(routes) else ""
            
            # Skip car rental entries
            if is_car_rental_row(rt):
                continue
            
            # Skip flight-only entries (Airlink, etc.) - they might be handled separately
            if "flight" in rt.lower() and "airport" not in rt.lower():
                continue
                
            sup_key = sup.lower().strip()
            
            if sup_key not in transfers_by_supplier:
                info = get_supplier_info(sup)
                transfers_by_supplier[sup_key] = TransferVoucher(
                    supplier=sup,
                    address=info.get("address", ""),
                    phone=info.get("phone", ""),
                    gps=info.get("gps", "")
                )
            
            # Parse pickup and dropoff from route
            pickup_loc = ""
            dropoff_loc = ""
            route_notes = ""
            
            if "-" in rt:
                parts = rt.split("-", 1)
                if len(parts) == 2:
                    first_part = parts[0].strip()
                    second_part = parts[1].strip()
                    
                    # Check if it's "Trf - Location" format
                    if first_part.lower() in ["trf", "transfer"]:
                        dropoff_loc = second_part
                    # Check if it has "incl." with additional info
                    elif "incl." in second_part.lower():
                        incl_parts = second_part.split("incl.", 1)
                        dropoff_loc = incl_parts[0].strip()
                        route_notes = "Includes: " + incl_parts[1].strip() if len(incl_parts) > 1 else ""
                        pickup_loc = first_part
                    else:
                        pickup_loc = first_part
                        dropoff_loc = second_part
            else:
                dropoff_loc = rt
            
            # Combine notes
            all_notes = []
            if route_notes:
                all_notes.append(route_notes)
            if row.get("transfer_notes"):
                all_notes.append(row.get("transfer_notes"))
            
            leg = TransferLeg(
                date=row["date"],
                pickup_location=pickup_loc,
                dropoff_location=dropoff_loc,
                pickup_time=pickup_times[idx] if idx < len(pickup_times) else "",
                dropoff_time=dropoff_times[idx] if idx < len(dropoff_times) else "",
                flight_number=flight_nums[idx] if idx < len(flight_nums) else "",
                notes="\n".join(all_notes)
            )
            transfers_by_supplier[sup_key].legs.append(leg)
    
    return list(transfers_by_supplier.values())


def parse_car_rentals(data_rows: List[Dict]) -> List[CarRentalVoucher]:
    """Parse car rental information from data rows."""
    car_rentals = []
    car_rental_data = {}
    
    for row in data_rows:
        route = row.get("transfer_route")
        if not route:
            continue
        
        routes = [r.strip() for r in route.split('\n') if r.strip()]
        
        for rt in routes:
            if is_car_rental_row(rt):
                # Extract car group info
                if "group" in rt.lower():
                    if "car_group" not in car_rental_data:
                        car_rental_data["car_group"] = rt
                        car_rental_data["pickup_date"] = row["date"]
                    car_rental_data["dropoff_date"] = row["date"]
                    
                if "collect" in rt.lower() or "pickup" in rt.lower():
                    car_rental_data["pickup_location"] = rt
                if "drop" in rt.lower() or "return" in rt.lower():
                    car_rental_data["dropoff_location"] = rt
    
    if car_rental_data.get("car_group"):
        info = get_supplier_info("pace car rental")  # Default car rental company
        
        # Parse car group for supplier info
        car_group = car_rental_data.get("car_group", "")
        
        car_rental = CarRentalVoucher(
            supplier="Pace Car Rental",  # Default, can be detected from ORGA
            car_group=car_group,
            pickup_date=car_rental_data.get("pickup_date"),
            pickup_location=car_rental_data.get("pickup_location", "Cape Town International Airport"),
            dropoff_date=car_rental_data.get("dropoff_date"),
            dropoff_location=car_rental_data.get("dropoff_location", "Cape Town International Airport"),
            address=info.get("address", ""),
            phone=info.get("phone", ""),
            gps=info.get("gps", "")
        )
        car_rentals.append(car_rental)
    
    return car_rentals


def parse_activities(data_rows: List[Dict]) -> List[ActivityVoucher]:
    """Parse activities from data rows, grouping by supplier."""
    activities_by_supplier: Dict[str, ActivityVoucher] = {}
    
    # Keywords that indicate restaurant (NOT activity)
    restaurant_only_keywords = ["dinner", "lunch"]
    
    # Keywords that indicate activity (even if at a restaurant-sounding venue)
    activity_keywords = ["tasting", "tour", "tickets", "watching", "drive", "panorama", "route", "safari"]
    
    for row in data_rows:
        supplier = row.get("activity_supplier")
        activity = row.get("activity_name")
        
        if not supplier:
            continue
        
        # Handle multi-line entries
        suppliers = [s.strip() for s in supplier.split('\n') if s.strip()]
        activities = [a.strip() for a in (activity or "").split('\n') if a.strip()]
        times = [t.strip() for t in (row.get("activity_time") or "").split('\n') if t.strip()]
        notes_list = [n.strip() for n in (row.get("activity_notes") or "").split('\n') if n.strip()]
        
        for idx, sup in enumerate(suppliers):
            act = activities[idx] if idx < len(activities) else ""
            time = times[idx] if idx < len(times) else ""
            notes = notes_list[idx] if idx < len(notes_list) else ""
            
            combined_text = (sup + " " + act + " " + notes).lower()
            
            # Check if this is explicitly an activity (e.g., wine tasting)
            is_explicit_activity = any(kw in combined_text for kw in activity_keywords)
            
            # Check if this is a restaurant meal
            is_restaurant_meal = any(kw in combined_text for kw in restaurant_only_keywords)
            
            # If it's a restaurant meal and NOT an activity, skip (handle in parse_restaurants)
            if is_restaurant_meal and not is_explicit_activity:
                continue
            
            # Check if this is a game drive at a safari lodge (should be part of hotel)
            if "game drive" in act.lower():
                continue  # Will be included in hotel voucher
            
            sup_key = sup.lower().strip()
            
            if sup_key not in activities_by_supplier:
                info = get_supplier_info(sup)
                activities_by_supplier[sup_key] = ActivityVoucher(
                    supplier=sup,
                    address=info.get("address", ""),
                    phone=info.get("phone", ""),
                    gps=info.get("gps", "")
                )
            
            entry = ActivityEntry(
                date=row["date"],
                activity_name=act or sup,  # Use supplier name if no activity specified
                time=time,
                notes=notes
            )
            activities_by_supplier[sup_key].entries.append(entry)
    
    return list(activities_by_supplier.values())


def parse_restaurants(data_rows: List[Dict]) -> List[RestaurantVoucher]:
    """Parse restaurant vouchers from data rows."""
    restaurants = []
    
    # Keywords for restaurant meals only
    restaurant_keywords = ["dinner", "lunch"]
    
    # Keywords that indicate activity (NOT restaurant) even at dining venues
    activity_keywords = ["tasting", "tour", "tickets", "watching"]
    
    for row in data_rows:
        supplier = row.get("activity_supplier")
        activity = row.get("activity_name")
        
        if not supplier:
            continue
        
        # Handle multi-line entries
        suppliers = [s.strip() for s in supplier.split('\n') if s.strip()]
        activities = [a.strip() for a in (activity or "").split('\n') if a.strip()]
        times = [t.strip() for t in (row.get("activity_time") or "").split('\n') if t.strip()]
        notes_list = [n.strip() for n in (row.get("activity_notes") or "").split('\n') if n.strip()]
        
        for idx, sup in enumerate(suppliers):
            act = activities[idx] if idx < len(activities) else ""
            time = times[idx] if idx < len(times) else ""
            notes = notes_list[idx] if idx < len(notes_list) else ""
            
            combined_text = (sup + " " + act + " " + notes).lower()
            
            # Check if this is a restaurant meal
            is_restaurant_meal = any(kw in combined_text for kw in restaurant_keywords)
            
            # Check if this is actually an activity (like wine tasting)
            is_activity = any(kw in combined_text for kw in activity_keywords)
            
            # Only include if it's a restaurant meal AND NOT an activity
            if not is_restaurant_meal or is_activity:
                continue
            
            info = get_supplier_info(sup)
            
            restaurant = RestaurantVoucher(
                supplier=sup,
                date=row["date"],
                time=time,
                notes=notes or act,  # Include activity description as notes
                address=info.get("address", ""),
                phone=info.get("phone", ""),
                gps=info.get("gps", "")
            )
            restaurants.append(restaurant)
    
    return restaurants


def parse_golf(data_rows: List[Dict]) -> List[GolfVoucher]:
    """Parse golf vouchers from data rows."""
    golf_vouchers = []
    
    for row in data_rows:
        supplier = row.get("golf_supplier")
        course = row.get("golf_course")
        
        if not supplier or not course:
            continue
        
        info = get_supplier_info(supplier)
        
        golf = GolfVoucher(
            supplier=supplier,
            course=course,
            date=row["date"],
            tee_time=row.get("tee_time", ""),
            cart=row.get("golf_cart", ""),
            rental_set=row.get("rental_set", ""),
            notes=row.get("golf_notes", ""),
            address=info.get("address", ""),
            phone=info.get("phone", ""),
            gps=info.get("gps", "")
        )
        golf_vouchers.append(golf)
    
    return golf_vouchers

